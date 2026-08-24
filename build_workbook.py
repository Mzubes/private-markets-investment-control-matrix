import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_competitive_matrix_workbook():
    """
    Reads matrix.csv and builds a clean Excel workbook with multiple sheets.
    Outputs: PrivateMarkets_InvestmentControl_CompetitiveMatrix.xlsx
    """
    
    # Read the CSV file
    df = pd.read_csv('matrix.csv')
    
    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    category_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    category_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Sheet 1: Full Matrix
    ws_full = wb.create_sheet("Full Matrix", 0)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_full.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            cell.border = border
    
    # Adjust column widths for Full Matrix
    ws_full.column_dimensions['A'].width = 18
    ws_full.column_dimensions['B'].width = 12
    ws_full.column_dimensions['C'].width = 18
    ws_full.column_dimensions['D'].width = 20
    ws_full.column_dimensions['E'].width = 25
    for col in range(6, len(df.columns) + 1):
        ws_full.column_dimensions[chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)].width = 14
    
    # Sheet 2: Company Overview
    ws_overview = wb.create_sheet("Company Overview", 1)
    overview_cols = ['Company', 'Founded', 'Funding if known', 'Primary customer', 'Core product']
    overview_df = df[overview_cols].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(overview_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_overview.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            cell.border = border
    
    ws_overview.column_dimensions['A'].width = 20
    ws_overview.column_dimensions['B'].width = 12
    ws_overview.column_dimensions['C'].width = 20
    ws_overview.column_dimensions['D'].width = 25
    ws_overview.column_dimensions['E'].width = 30
    
    # Sheet 3: Data Capabilities
    ws_data = wb.create_sheet("Data Capabilities", 2)
    data_cols = ['Company', 'Document collection', 'Document extraction', 'LPA extraction', 
                 'Side-letter extraction', 'Investor-specific term modeling']
    data_df = df[data_cols].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            cell.border = border
    
    ws_data.column_dimensions['A'].width = 20
    for col in range(2, 7):
        ws_data.column_dimensions[chr(64 + col)].width = 16
    
    # Sheet 4: Workflow & Automation
    ws_workflow = wb.create_sheet("Workflow & Automation", 3)
    workflow_cols = ['Company', 'Capital-call extraction', 'Capital-call validation', 
                     'Distribution validation', 'Fee verification', 'Carry verification', 
                     'Expense verification', 'Human approval workflows', 'Downstream execution']
    workflow_df = df[workflow_cols].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(workflow_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_workflow.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            cell.border = border
    
    ws_workflow.column_dimensions['A'].width = 20
    for col in range(2, 10):
        ws_workflow.column_dimensions[chr(64 + col)].width = 15
    
    # Sheet 5: Reconciliation & Monitoring
    ws_recon = wb.create_sheet("Reconciliation & Monitoring", 4)
    recon_cols = ['Company', 'NAV reconciliation', 'Cash-flow reconciliation', 
                  'Expected-vs-actual calculations', 'Exception detection', 
                  'Exception investigation', 'Portfolio monitoring', 
                  'Underwriting-vs-actual monitoring']
    recon_df = df[recon_cols].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(recon_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_recon.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            cell.border = border
    
    ws_recon.column_dimensions['A'].width = 20
    for col in range(2, 9):
        ws_recon.column_dimensions[chr(64 + col)].width = 18
    
    # Sheet 6: Integration & Support
    ws_integration = wb.create_sheet("Integration & Support", 5)
    integration_cols = ['Company', 'Accounting integrations', 'Portfolio-system integrations', 
                        'API availability', 'Audit trail', 'Evidence/provenance', 
                        'Pre-investment functionality', 'Post-investment functionality',
                        'Customer/VPC deployment', 'Known pricing', 'Notable customers']
    integration_df = df[integration_cols].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(integration_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_integration.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            cell.border = border
    
    ws_integration.column_dimensions['A'].width = 20
    for col in range(2, 12):
        ws_integration.column_dimensions[chr(64 + col)].width = 16
    
    # Save the workbook
    wb.save('PrivateMarkets_InvestmentControl_CompetitiveMatrix.xlsx')
    print("Workbook created successfully: PrivateMarkets_InvestmentControl_CompetitiveMatrix.xlsx")

if __name__ == "__main__":
    create_competitive_matrix_workbook()
